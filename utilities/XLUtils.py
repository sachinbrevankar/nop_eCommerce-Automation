from openpyxl import load_workbook

class XLUtils:
    @staticmethod
    def getRowCount(file, sheetName):
        workbook = load_workbook(file)
        sheet = workbook[sheetName]
        return sheet.max_row

    @staticmethod
    def getColumnCount(file, sheetName):
        workbook = load_workbook(file)
        sheet = workbook[sheetName]
        return sheet.max_column

    @staticmethod
    def readData(file, sheetName, rownum, columnno):
        workbook = load_workbook(file)
        sheet = workbook[sheetName]
        return sheet.cell(row=rownum, column=columnno).value

    @staticmethod
    def writeData(file, sheetName, rownum, columnno, data):
        workbook = load_workbook(file)
        sheet = workbook[sheetName]
        sheet.cell(row=rownum, column=columnno).value = data
        workbook.save(file)
